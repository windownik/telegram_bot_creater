import openpyxl


# Создаем список хэндлеров
def create_list():
    excel = openpyxl.reader.excel.load_workbook(filename='creator.xlsx')
    excel.active = 0
    sheet = excel.active
    i = 2
    handlers_text = []
    proces = True
    while proces:
        if sheet[f'B{i}'].value is not None:
            pre_text = f'{i} |--| '
            text = str(sheet[f'B{i}'].value)
            handlers_text.append((pre_text + text))
            i += 1
        else:
            proces = False

    return handlers_text


# Сохраняем данные хэндлера
def change_item(index: str, main_text: str, previos_state: str, next_state: str):
    excel = openpyxl.reader.excel.load_workbook(filename='creator.xlsx')
    excel.active = 0
    sheet = excel.active
    sheet[f'A{index}'] = previos_state
    sheet[f'B{index}'] = main_text
    sheet[f'C{index}'] = next_state
    excel.save('creator.xlsx')


# Получаем данные по стэйтам
def read_item_data(index: str):
    excel = openpyxl.reader.excel.load_workbook(filename='creator.xlsx')
    excel.active = 0
    sheet = excel.active
    pre_states = str(sheet[f'A{index}'].value).split('###')
    next_states = str(sheet[f'C{index}'].value).split('###')
    pre_states = ','.join(str(i) for i in pre_states if i is not None)
    next_states = ','.join(str(i) for i in next_states if i is not None)
    return (pre_states, next_states)

