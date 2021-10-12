import os
import openpyxl


# Создаем main.py
def creat_main():
    try:
        file = open('bot/main.py', 'w')
    except:
        os.mkdir("bot")
        file = open('bot/main.py', 'w')
    text = """
from aiogram import executor
from modules.dispatcher import dp
import __init__


if __name__ == '__main__':
    executor.start_polling(dp)
        """
    file.write(text)
    file.close()


# Создаем файл настроек телеграма
def creat_json():
    try:
        file = open('bot/telegram.json', 'w')
    except:
        os.mkdir("bot")
        file = open('bot/telegram.json', 'w')
    text = """
{
  "telegram_token": "1849931200:AAE1eurg0RIgrC2pU55WpLIi8diXiV-Jei8"
}
        """
    file.write(text)
    file.close()


# Создаем файл диспетчера
def creat_dispatcher():
    try:
        file = open('bot/modules/dispatcher.py', 'w')
    except:
        os.mkdir("bot/modules")
        file = open('bot/modules/dispatcher.py', 'w')
    text = """
from aiogram import Bot, Dispatcher
from aiogram.contrib.fsm_storage.memory import MemoryStorage

from modules.states import AllStates
import json


import logging

with open("telegram.json", 'r') as file:
    telegram_token = str(json.load(file)["telegram_token"])
    file.close()

storage = MemoryStorage()
logging.basicConfig(level=logging.INFO)
bot = Bot(telegram_token)
dp = Dispatcher(bot, storage=storage)

        """
    file.write(text)
    file.close()


# Создаем state
def creat_states():
    try:
        file = open('bot/modules/states.py', 'w')
    except:
        os.mkdir("bot/modules")
        file = open('bot/modules/states.py', 'w')
    excel = openpyxl.reader.excel.load_workbook(filename='creator.xlsx')
    excel.active = 0
    sheet = excel.active
    i = 2
    questions_text = ''
    proces = True
    while proces:
        if sheet[f'B{i}'].value is not None:
            questions_text = questions_text +  f"""
    question_{i} = State()"""
            i += 1
        else:
            proces = False

    text = """
from aiogram.dispatcher.filters.state import State, StatesGroup


# Welcome form
class AllStates(StatesGroup):
        """
    file.write(text + questions_text)
    file.close()


# Создаем __init__
def creat_init():
    try:
        file = open('bot/__init__.py', 'w')
    except:
        os.mkdir("bot")
        file = open('bot/__init__.py', 'w')
    text = """
from handlers import start_handlers

        """
    file.write(text)
    file.close()


# Создаем клавиатуры
def creat_keyboards():
    try:
        file = open('bot/modules/keybords.py', 'w', encoding='utf-8')
    except:
        os.mkdir("bot")
        os.mkdir("modules")
        file = open('bot/modules/keybords.py', 'w', encoding='utf-8')
    start_text = """
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
"""
    excel = openpyxl.reader.excel.load_workbook(filename='creator.xlsx')
    excel.active = 0
    sheet = excel.active
    i = 2
    new_kb_text = ''
    btn_text = ''
    btn_add_text = ''
    proces = True
    while proces:
        if "reply" in str(sheet[f'D{i}'].value):
            new_kb_text = new_kb_text + f'''
keyboard_{i} = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)'''
            letters = ("E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O")
            for l in letters:
                if sheet[f'{l}{i}'].value is not None:
                    btn_text = btn_text + f"""                   
btn_{i}_{l} = KeyboardButton('{str(sheet[f'{l}{i}'].value)}')"""
                    btn_add_text = btn_add_text + f"""    
keyboard_{i}.add(btn_{i}_{l})"""
            i += 1
        elif "inline" in str(sheet[f'D{i}'].value):
            new_kb_text = new_kb_text + f'''
keyboard_{i} = InlineKeyboardMarkup()'''
            letters = ("E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O")
            for l in letters:
                if sheet[f'{l}{i}'].value is not None:
                    btn_text = btn_text + f"""                   
btn_{i}_{l} = InlineKeyboardButton(text='{str(sheet[f'{l}{i}'].value).split("###")[0]}', callback_data='{str(sheet[f'{l}{i}'].value).split("###")[1]}')"""
                    btn_add_text = btn_add_text + f"""    
keyboard_{i}.add(btn_{i}_{l})"""
            i += 1
        else:
            proces = False
    text = start_text + "\n" + new_kb_text + "\n" + btn_text +"\n" + btn_add_text
    file.write(text)
    file.close()


# Создаем хэндлеры
def creat_hendler():
    try:
        file = open('bot/handlers/start_handlers.py', 'w', encoding='utf-8')
    except:
        os.mkdir("bot/handlers")
        file = open('bot/handlers/start_handlers.py', 'w', encoding='utf-8')
    text_standart = """
from aiogram import types
from main import dp
from aiogram.dispatcher.filters import Text
import logging
from aiogram.dispatcher import FSMContext
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from modules.states import AllStates
from modules import keybords


# Start menu
@dp.message_handler(commands=['start'], state='*')
async def start_menu(message: types.Message):
    await message.answer(text='Привет! Ты попал в Телеграм бот.\\n'
                              'Для получения общей информации о нашей деятельности нажми /help\\n'
                              'Для отмены всех действий в любой момент нажмите /cancel')
    await AllStates.question_2.set()
    
        """
    excel = openpyxl.reader.excel.load_workbook(filename='creator.xlsx')
    excel.active = 0
    sheet = excel.active
    i = 2
    handlers_text = ''
    proces = True
    while proces:
        if sheet[f'B{i}'].value is not None:
            text = str(sheet[f'B{i}'].value)
            if "reply" in str(sheet[f'D{i}'].value):
                handlers_text = handlers_text + f"""
    
# Question handler №{i}
@dp.message_handler(state=AllStates.question_{i})
async def question_{i}_menu(message: types.Message):
    await message.answer(text='{text}', reply_markup=keybords.keyboard_{i})
    """
            elif "inline" in str(sheet[f'D{i}'].value):
                handlers_text = handlers_text + f"""

# Question handler №{i}
@dp.message_handler(state=AllStates.question_{i})
async def question_{i}_menu(message: types.Message):
    await message.answer(text='{text}', reply_markup=keybords.keyboard_{i})
    """
            else:
                handlers_text = handlers_text + f"""

# Question handler №{i}
@dp.message_handler(state=AllStates.question_{i})
async def question_{i}_menu(message: types.Message):
    await message.answer(text='{text}', 
    reply_markup=types.ReplyKeyboardRemove())
"""
            if sheet[f'B{i+1}'].value is not None:
                handlers_text = handlers_text + f"""
    await AllStates.question_{i + 1}.set()
                    """
            else:
                pass
            i += 1
        else:
            proces = False

    file.write(text_standart + handlers_text)
    file.close()
